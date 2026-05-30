"""
Standalone batch processing module for running cell segmentation
on remote servers without the Streamlit UI.

Usage (CLI):
    python batch_processor.py --input-dir /path/to/images --output /path/to/results

Usage (Python):
    from batch_processor import process_batch
    zip_path, stats_df = process_batch("/path/to/images", "/path/to/results")
"""

import os
import io
import math
import zipfile
import shutil
import tempfile
import argparse
import logging
from time import time

import pandas as pd
import matplotlib.pyplot as plt
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from core import process_image_from_path
from utils import (
    sort_images_incucyte,
    extract_incucyte_info,
    combine_image_statistics,
    empty_image_statistics,
)
from core.image_processing import get_image_files

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


def process_batch(
    input_source,
    output_dir=None,
    model_type="vit_b_lm",
    min_area=200,
    numbered=False,
):
    """
    Process a batch of images and save results (CSVs, visualizations, ZIP).

    Args:
        input_source: Directory path containing images, or a list of image file paths.
        output_dir: Directory to save outputs. If None, uses a temp directory.
        model_type: MicroSAM model type.
        min_area: Minimum area threshold for cell filtering.
        numbered: Write cell number on each cell.

    Returns:
        Tuple of (zip_path, stats_df) where zip_path is the path to the
        results ZIP archive and stats_df contains combined statistics.
        Returns (None, None) if no images were processed successfully.
    """
    # Resolve image paths
    if isinstance(input_source, str):
        image_paths = get_image_files(input_source)
        if not image_paths:
            logger.error("No images found in %s", input_source)
            return None, None
    else:
        image_paths = list(input_source)

    # Create output directory
    if output_dir is None:
        output_dir = tempfile.mkdtemp(prefix="batch_results_")
    else:
        os.makedirs(output_dir, exist_ok=True)

    logger.info("Found %d images", len(image_paths))
    logger.info("Model: %s | Min area: %d | Numbered: %s", model_type, min_area, numbered)

    # Group images by incucyte metadata
    incucyte_group = sort_images_incucyte(
        images=[os.path.basename(p) for p in image_paths]
    )

    all_image_stats = []
    successful = 0
    failed = 0

    for idx, image_path in enumerate(tqdm(image_paths, desc="Processing", unit="image")):
        filename = os.path.splitext(os.path.basename(image_path))[0]
        logger.info("Processing %s (%d/%d)", filename, idx + 1, len(image_paths))

        # Process the image using the shared core logic
        try:
            result = process_image_from_path(image_path, model_type, min_area, numbered)
        except Exception as e:
            logger.warning("Exception processing %s: %s", filename, e)
            result = None

        if result is None:
            logger.warning("Failed to process %s, recording zeros", filename)
            failed += 1
            incucyte_info = extract_incucyte_info(os.path.basename(image_path))
            all_image_stats.append(empty_image_statistics(filename))
            continue

        segmentation = result["segmentation"]
        features_df = result["features_df"]
        image_stats = result["image_stats"]
        visualizations = result["visualizations"]
        incucyte_info = result["incucyte_info"]

        # Track incucyte grouping
        incucyte_group[incucyte_info["key"]]["results"].append(
            tuple([incucyte_info["position"], image_stats, features_df, segmentation])
        )
        all_image_stats.append(image_stats)

        # Save features CSV
        features_csv_path = os.path.join(output_dir, f"{filename}_features.csv")
        features_df.to_csv(features_csv_path, index=False)

        # Create and save 3-panel visualization
        vis_output_path = os.path.join(output_dir, f"{filename}_visualization.png")
        _save_visualization(
            result["image"],
            visualizations["final_filtered_vis"],
            visualizations["segmentation_vis"],
            len(features_df),
            int(segmentation.max()),
            vis_output_path,
        )

        successful += 1

    logger.info("Done. %d succeeded, %d failed.", successful, failed)

    if not all_image_stats:
        logger.warning("No images were successfully processed.")
        return None, None

    # Build combined statistics Excel file with embedded graphs
    stats_path = _build_excel_with_graphs(all_image_stats, output_dir)

    # Create ZIP archive
    zip_path = _create_results_zip(output_dir)

    # Load stats DataFrame for return value
    stats_df = pd.read_excel(stats_path, sheet_name="data")

    logger.info("Results saved to %s", zip_path)
    return zip_path, stats_df


def _save_visualization(original_image, final_filtered_vis, segmentation_vis,
                        num_filtered, num_all, output_path):
    """Save a 3-panel comparison visualization."""
    fig, axes = plt.subplots(1, 3, figsize=(20, 5))

    axes[0].imshow(original_image)
    axes[0].set_title("Original Image")
    axes[0].axis("off")

    axes[1].imshow(final_filtered_vis)
    axes[1].set_title(f"Final Filtered ({num_filtered} cells)")
    axes[1].axis("off")

    axes[2].imshow(segmentation_vis)
    axes[2].set_title(f"All Cells ({num_all})")
    axes[2].axis("off")

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()


def _build_excel_with_graphs(all_image_stats, output_dir):
    """Create FINAL_STATS.xlsx with a data sheet and embedded graphs."""
    final_combined_stats = combine_image_statistics(all_image_stats)
    stats_df = pd.DataFrame(final_combined_stats)

    xlsx_path = os.path.join(output_dir, "FINAL_STATS.xlsx")
    stats_df.to_excel(xlsx_path, index=False, sheet_name="data")

    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("graphs")

    df_sorted = stats_df.sort_values("image_name")
    numeric_cols = df_sorted.select_dtypes(include="number").columns

    row = 1
    for col in numeric_cols:
        plt.figure(figsize=(8, 5))
        plt.plot(df_sorted["image_name"], df_sorted[col], marker="o")
        plt.title(col)
        plt.xticks(rotation=45)
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format="png")
        plt.close()
        buf.seek(0)

        img = XLImage(buf)
        img.width, img.height = 640, 400
        ws.add_image(img, f"A{row}")
        row += 20

    wb.save(xlsx_path)
    logger.info("Saved combined statistics to %s", xlsx_path)
    return xlsx_path


def _create_results_zip(output_dir):
    """Create a ZIP archive of all results in output_dir."""
    zip_path = os.path.join(output_dir, "results.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(output_dir):
            for file in files:
                if file == "results.zip":
                    continue  # don't nest the zip in itself
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, output_dir)
                zipf.write(file_path, arcname)

    logger.info("Created ZIP archive: %s", zip_path)
    return zip_path


def main():
    """CLI entry point for remote batch processing."""
    parser = argparse.ArgumentParser(
        description="Batch process images with MicroSAM cell segmentation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python batch_processor.py --input-dir ./images --output ./results\n"
            "  python batch_processor.py --input-dir ./images --min-area 300 --numbered\n"
        ),
    )
    parser.add_argument(
        "--input-dir",
        type=str,
        required=True,
        help="Directory containing input images",
    )
    parser.add_argument(
        "--output",
        type=str,
        required=True,
        help="Directory to save results (CSVs, visualizations, ZIP)",
    )
    parser.add_argument(
        "--model",
        type=str,
        default="vit_b_lm",
        help="MicroSAM model type (default: vit_b_lm)",
    )
    parser.add_argument(
        "--min-area",
        type=int,
        default=500,
        help="Minimum cell area in pixels² for filtering (default: 500)",
    )
    parser.add_argument(
        "--numbered",
        action="store_true",
        help="Draw cell numbers on visualization images",
    )

    args = parser.parse_args()

    start = time()
    zip_path, stats_df = process_batch(
        input_source=args.input_dir,
        output_dir=args.output,
        model_type=args.model,
        min_area=args.min_area,
        numbered=args.numbered,
    )
    elapsed = time() - start

    if zip_path and stats_df is not None:
        print(f"\nCompleted in {elapsed / 60:.2f} minutes.")
        print(f"Total images processed: {len(stats_df)}")
        print(f"Results ZIP: {zip_path}")
    else:
        print("\nNo images were processed successfully.")
        exit(1)


if __name__ == "__main__":
    main()
